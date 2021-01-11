# Adding images into excel sheet using openpyxl

from openpyxl import Workbook
from openpyxl.drawing.image import Image

from openpyxl.utils import FORMULAE
#********************************************* add images ************************************************
work_book = Workbook()
sheet =  work_book.active()
img = Image('img_path')
img.width = 10
img.height = 10
sheet.add_image(img, 'C1') # embedding image into excel
work_book.save('file_path')

#********************************************* use formulas for cell value calculation ************************************************
sheet['C2'] = 'sum: '
sheet['D2'] = '=SUM(A1:A5)'

sheet['C3'] = 'Product: '
sheet['D3'] = '=PRODUCT(A1:A5)'

sheet['C4'] = 'Average: '
sheet['D4'] = '=AVERAGE(A1:A5)'

sheet['C5'] = 'COUNT: '
sheet['D5'] = '=COUNT(A1:A5)'

work_book.save('file_path')

#********************************************* implement formulas for cell value calculation ************************************************
header = ['Cake','Quantity','Price','Revenue']
data = [['Chocolate',30,3],
		 ['CheeseCake',50,10],
		 ['Carrot',25,10],
		 ['Milk',20,50],
		 ['Red Velvet',200,4]
		]
work_book.create_sheet('CakeSales',index=0)
cake_sales_sheet = work_book['CakeSales']

cake_sales_sheet.append(header)
for row in data:
	cake_sales_sheet.append(row)
max_row_str = str(cake_sales_sheet.max_row)

for row in cake_sales_sheet['D2:D'+ max_row_str]:
	for cell in row:
		cell.value = '=$B${0}*$C${0}'.format(cell.row)
		
total_row_str = str(cake_sales_sheet.max_row+2)
cake_sales_sheet['C'+total_row_str] = 'Total Sales : '
cake_sales_sheet['D'+ total_row_str] = '=SUM(D2:D'+max_row_str+')'
work_book.save('file_path')

for row in cake_sales_sheet['C2:D'+max_row_str]:
	for cell in row:
		cell.number_format = '$#,##0.00' # giving $ symbol and other formats to product values
cake_sales_sheet['D'+ total_row_str].number_format = '$#.##0.00'  # giving $ symbol and other formats to sum value
work_book.save('file_path')
	
	
	
	
	
	
	
	
	
	
	
	
	
	
	
	
extended iterable unpacking 