# Implementing VLOOKUPS in excel using openpyxl module. namedRange is the collection of individual cells 
import openpyxl

work_book = openpyxl.load_workbook('file_path') # this file consists of 2 sheets. One is product and the other is data.
sheet = work_book['Products']  # a named-range(table) is defined in products sheet and another one is there in the other sheet
fx_rate = work_book.defined_names['fx_rate']  # creating an object of named-range(table)

cells =[]
for title, coord in fx_rate.destinations:
	ws = work_book[title]  # title means title of sheet.
	cells.append(ws[coord])  # this will append the cell name like A1 or B1. 

max_row_str = str(sheet.max_row)
for row in sheet['C3:C'+ max_row_str]:
	for cell in row:
		cell.value = '=$B${0}*VLOOKUP($C$2, fx_rate, 2, False)'.format(cell.row) # =VLOOKUP (value, table, col_index, [range_lookup])
		cell.number_format = '#,##0.00'
work_book.create_named_range('Products',sheet,'$A$3:$B$'+max_row_str)
work_book.save('file_path')
