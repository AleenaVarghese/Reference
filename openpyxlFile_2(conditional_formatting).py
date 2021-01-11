# Font Property of Excel using openpyxl 
# conditional formating of rows using openpyxl

import openpyxl
from openpyxl import Font, Color, Alignment, Border, Side, Colors
from openpyxl import NamedStyle
from openpyxl.styles.fills impprt PatternFill

from openpyxl.styles import PatternFill, Colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.formatting.rule import IconSetRule
from openpyxl.formatting.rule import dataBarRule

work_book = openpyxl.load_workbook('path')
sheet = work_book['sheet']
#********************************************* give styles individually ************************************************
 
bold_font = Font(bold =True)
big_red_text = Font(color = colors.RED, size = 20)
center_aligned_text =  Alignment(horizontal = 'center')
double_border_side = side(border_style='double')
squre_border = Border(top = double_border_side,
					  right = double_border_side,
					  bottom = double_border_side,
					  left = double_border_side)
					  
sheet['B2'].font = bold_font
sheet['B3'].font = big_red_text
sheet['C2'].alignment = center_aligned_text
sheet['C3'].border = squre_border
#********************************************* give styles to a collection of cells or rows ************************************************

custom_style1 = NamedStyle(name='header')
custom_style1.font = Font(bold =  True)
custom_style1.border = Border(bottom = side(border_style = 'thin'))
custom_style1.alignment = Alignment(horizontal ='center', vertical = 'center')
header_row = sheet[1]  # handling first row of the sheet.

for cell in header_row:
	cell.style = custom_style1
work_book.save('path')

custom_style2 = NamedStyle(name = 'highlight')
custom_style2.fill = PatternFill(fgColor = colors.color('d7abcc'), 
								 patterntype = 'lightHorizontal') # to set foreground color to specified areas
for cell in sheet['A']: # Will iterate over all cells in column A which contains some data.
	cell.style = custom_style2
work_book.save('path')

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ Conditional Formating  @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

#********************************************* setting the color of rows based on condition  ************************************************

yellow_background = PatternFill(bgColor = colors.YELLOW)
diff_style = DifferentialStyle(fill = yellow_background)
rule = Rule(type ='expression', dxf = diff_style)
rule.formula = ['$M1<70000']
sheet.conditional_formating.add(sheet.calculate_dimension(), rule) # important function and has two parameters first one is data range and then rule to be applied.
work_book.save(path)

#********************************************* Applying Conditional Formating using ColorScaleRule  ************************************************

color_scale_rule1 = ColorScaleRule(start_type = 'min',start_color = colors.YELLOW,
								   end_type = 'max', end_color = colors.RED)
sheet.conditional_formating.add(sheet.calculate_dimension() ,color_scale_rule1)								  
work_book.save(path)
color_scale_rule2 = ColorScaleRule(start_type = 'percentile', start_value = 0, start_color = colors.GREEN,
								   mid_type = 'percentile', mid_value = 50, mid_color = colors.	RED,
								   end_type = 'percentile', end_value = 90,end_color = colors.BLUE)
sheet.conditional_formating.add(sheet.calculate_dimension(), color_scale_rule2)
sheet.save(path)

#********************************************* Applying Conditional Formating using IconSetRule  ************************************************
# will show icons along with column values based on the value.
icon_set_rule = IconSetRule(icon_style = '4arrows',
							type = 'num',
							values = [1,2,3,4,5])
sheet.conditional_formating.add(sheet.calculate_dimension(), icon_set_rule) 
sheet.save(path)

#********************************************* Applying Conditional Formating using DataBarRule  ************************************************
# will show bars along with column values based on the value.
data_bar_rule = DataBarRule(start_type = 'num', start_value = '1', 
							end_type = 'num', end_value = '4',
							color = colors.RED)
sheet.conditional_formating.add(sheet.calculate_dimension(), data_bar_rule)
sheet.save(path)


